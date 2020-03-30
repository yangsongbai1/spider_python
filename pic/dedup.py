import openpyxl
import tldextract
import requests
from lxml import etree
import re


def dedup():
	work = openpyxl.load_workbook('医院网站采集--王彦超.xlsx')
	sheet = work.active

	new_work = openpyxl.Workbook()
	new_sheet = new_work.active


	for row in sheet:
		line = [i.value for i in row]
		url = line[2]
		tld = tldextract.extract(url)
		fqdn = tld.fqdn
		url = 'http://' + fqdn
		line[2] = url
		new_sheet.append(line)

	new_work.save('医院网站采集--王彦超11.xlsx')

def dowload(url):
	try:
		resp = requests.get(url, timeout=5)
	except:
		return ''
	try:
		charset = re.search('''charset\s*=(.*?)\>''', resp.text).group(1)
		if 'gb' in charset.lower():
			charset = 'gbk'
		else:
			charset = 'utf-8'
	except:
		charset = 'utf-8'
	resp.encoding = charset
	resp = resp.text
	try:
		title = re.search('''\<title\>(.*?)\</title\>''', resp).group(1)
		title = title.split('-')[0].strip()
	except:
		return ''
	return title

def get_title():
	work = openpyxl.load_workbook('恒大健康.xlsx')
	sheet = work.active

	new_work = openpyxl.Workbook()
	new_sheet = new_work.active

	for row in sheet:
		line = [i.value for i in row]
		url = line[1]
		title = dowload(url)
		if title:
			line[0] = title
		new_sheet.append(line)
		print(line)
	new_work.save('恒大健康ss.xlsx')

dedup()